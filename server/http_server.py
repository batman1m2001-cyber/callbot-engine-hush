"""HTTP server — health, customer-info, call summary endpoints.

Port: 9923 (configurable via HTTP_API_PORT env var).
"""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, UploadFile, File
from pydantic import BaseModel

from server import config, customer_store

LOGGER = logging.getLogger(__name__)

app = FastAPI(title="Callbot Engine Hush — HTTP API")


# ── Models ───────────────────────────────────────────────────────────────────

class CustomerInfo(BaseModel):
    Customer_id: str
    student_name: str = ""
    class_time: str = ""
    program_name: str = ""
    hotline: str = ""
    agent_name: str = ""
    phone_number: str = ""
    # Allow extra fields
    model_config = {"extra": "allow"}


class CallSummaryRequest(BaseModel):
    call_ids: List[str] = []
    date: Optional[str] = None  # YYYY-MM-DD


# ── Health ───────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"status": "healthy", "redis_status": "n/a", "version": "hush-1.0"}


@app.get("/version")
async def version():
    return {"version": "hush-1.0", "engine": "callbot-engine-hush"}


# ── Customer Info ────────────────────────────────────────────────────────────

@app.post("/api/v1/customer-info")
async def post_customer_info(info: CustomerInfo):
    """Store script_data for a customer."""
    data = info.model_dump()
    customer_id = data.pop("Customer_id")
    data["Customer_id"] = customer_id  # keep in data too
    customer_store.save(customer_id, data)
    return {"status": "ok", "customer_id": customer_id}


@app.post("/api/v1/customer-info/upload-xlsx")
async def upload_customer_xlsx(file: UploadFile = File(...)):
    """Bulk upload customer info from Excel file."""
    try:
        import openpyxl
    except ImportError:
        return {"status": "error", "message": "openpyxl not installed"}

    content = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        customer_id = str(data.get("Customer_id", ""))
        if not customer_id:
            continue
        customer_store.save(customer_id, data)
        count += 1

    return {"status": "ok", "imported": count}


# ── Call Summary ─────────────────────────────────────────────────────────────

@app.post("/api/v1/calls/summary/batch")
async def call_summary_batch(req: CallSummaryRequest):
    """Retrieve call logs by call_ids and/or date."""
    results = []

    if req.date:
        # Parse date → YYYYMMDD
        try:
            dt = datetime.strptime(req.date, "%Y-%m-%d")
            date_dir = dt.strftime("%Y%m%d")
        except ValueError:
            return {"status": "error", "message": f"Invalid date format: {req.date}"}

        jsonl_path = Path(config.LOG_BASE_DIR) / "educa_reminder" / date_dir / "calls.jsonl"
        if jsonl_path.exists():
            with open(jsonl_path) as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        entry = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    # Filter by call_ids if provided
                    if req.call_ids and entry.get("call_id") not in req.call_ids:
                        continue
                    results.append(entry)
    elif req.call_ids:
        # Search recent logs (last 7 days)
        base = Path(config.LOG_BASE_DIR) / "educa_reminder"
        if base.exists():
            for date_dir in sorted(base.iterdir(), reverse=True)[:7]:
                jsonl_path = date_dir / "calls.jsonl"
                if not jsonl_path.exists():
                    continue
                with open(jsonl_path) as f:
                    for line in f:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            entry = json.loads(line)
                        except json.JSONDecodeError:
                            continue
                        if entry.get("call_id") in req.call_ids:
                            results.append(entry)

    return {"status": "ok", "count": len(results), "calls": results}
